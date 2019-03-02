from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

#FFFF8080 - выходной
#FFCC00FF - больничный
#FF66FFFF - отпуск
#FF00B050 - замена
#Доп.Часы - FF0070C0 (версия рыжего), нормальная - "4", лол

#Собрать имена всех ноков и номер строки, в котором она содержится
def get_noc_names(sheet):
	names = {}
	for cellObj in sheet['A5':'A40']:
		for cell in cellObj:
			if cell.value is None:
				continue
			elif 'Команда' in cell.value:
				continue
			else: names[cell.value] = cell.coordinate
	return names

#Получаем номер строки с первой функции
def get_noc_row(noc_name, sheet):
	names = get_noc_names(sheet)
	row = names[noc_name]
	return row

#Собираем список словарей с сменой и цветом ячейки
def get_shifts(noc_name, start, end, sheet):
	row = get_noc_row(noc_name, sheet)
	shifts = []
	noc_start = start + row[1:]
	noc_end = end + row[1:]
	for cellObj in sheet[noc_start:noc_end]:
		for cell in cellObj:
			try:
				shifts.append({'data': {'shift': cell.value, 'color': cell.fill.start_color.index}})
			except:
				shifts.append({'data': {'shift': cell.value, 'color': "None"}})

	return shifts

#Стата для одного человека
def get_stats(noc_name, start, end, sheet):
	shifts = get_shifts(noc_name, start, end, sheet)
	total_hours = 0
	day_shifts = 0
	night_shifts = 0
	vacations = 0
	noc_hours = 0
	for things in shifts:
		shift = things['data']['shift']
		color = things['data']['color']
		print (noc_name)
		print (color)
		
		if things['data']['color'] == 4:
			print ("ok, it found it was 4")
			noc_hours += shift
			continue
		elif color == "FF0070C0": #Рыжий Богдан, если ты читаешь этот код, то знай
			noc_hours += shift #Это ебучее условие было написано потому-что ты путаешь цвета
			continue
		elif shift is None:
			continue
		else:
			for item in str(shift):
				if any ([item == '1', item == '2']):
					day_shifts += 1
					total_hours += 8
				elif item == '3':
					night_shifts += 1
					total_hours += 16

	return total_hours, day_shifts, night_shifts, noc_hours

#Цикл для сбора инфы по всем
def get_all_stats(start, end, files, end1):
	new_wb = Workbook()
	new_sheet = new_wb.active
	if files == 2:
		file = 'stats/upload/noc_schedule2.xlsx'
		wb = load_workbook(file)
		sheet = wb['Sheet1']
	else:
		file = 'stats/upload/noc_schedule1.xlsx'
		wb = load_workbook(file)
		sheet = wb['Sheet1']

	names = get_noc_names(sheet)
	new_sheet['A1'] = "Names"
	new_sheet['B1'] = "Day Shifts"
	new_sheet['C1'] = "Night Shifts"
	new_sheet['D1'] = "Noc Hours"
	new_sheet['E1'] = "Total Hours"
	i=1
	if files == 1:
		for name in names:
			i+=1
			total_hours, day_shifts, night_shifts, noc_hours = get_stats(name, start, end, sheet)
			new_sheet['A'+str(i)] = name
			new_sheet['B'+str(i)] = day_shifts
			new_sheet['C'+str(i)] = night_shifts
			new_sheet['D'+str(i)] = noc_hours
			new_sheet['E'+str(i)] = total_hours
			new_wb.save("stats/upload/noc_schedule_stats.xlsx")
	if files == 2:
		for name in names:
			i+=1
			file = 'stats/upload/noc_schedule1.xlsx'
			wb = load_workbook(file)
			sheet = wb['Sheet1']
			total_hours1, day_shifts1, night_shifts1, noc_hours1 = get_stats(name, start, end, sheet)

			file = 'stats/upload/noc_schedule2.xlsx'
			wb = load_workbook(file)
			sheet = wb['Sheet1']
			total_hours2, day_shifts2, night_shifts2, noc_hours2 = get_stats(name, "A", end1, sheet)

			total_hours = total_hours1 + total_hours2
			day_shifts = day_shifts1 + day_shifts2
			night_shifts = night_shifts1 + night_shifts2
			noc_hours = noc_hours1 + noc_hours2
			new_sheet['A'+str(i)] = name
			new_sheet['B'+str(i)] = day_shifts
			new_sheet['C'+str(i)] = night_shifts
			new_sheet['D'+str(i)] = noc_hours
			new_sheet['E'+str(i)] = total_hours
			new_wb.save("stats/upload/noc_schedule_stats.xlsx")
